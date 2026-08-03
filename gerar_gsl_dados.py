#!/usr/bin/env python3
"""
GSL Bartofil v7 — gerador da planilha GSL-DADOS (camada restrita).

Fonte da verdade da estrutura: este arquivo. A planilha e um artefato de build.
Uso:  python3 build/gerar_gsl_dados.py [saida.xlsx]

Regras de projeto respeitadas aqui:
  - abas de dados sao PLANAS: cabecalho na linha 1, sem titulos mesclados,
    sem linha em branco. E o que o Looker Studio e o Apps Script esperam.
  - nenhuma formula: todo calculo e feito pelo Apps Script e gravado como valor.
    Isso evita limite de linhas com formula e mantem o Looker rapido.
  - abas de configuracao (FONTES, PARAM) sao localizadas pelo script por
    MARCADOR na coluna A, nunca por numero fixo de linha.
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

AZUL = "111785"
VERDE = "01973A"
AMARELO = "FFEE03"
CINZA = "F1EFE8"
CINZA_ESC = "5F5E5A"
BRANCO = "FFFFFF"

FONTE_PADRAO = "Arial"

hdr_font = Font(name=FONTE_PADRAO, size=10, bold=True, color=BRANCO)
hdr_fill = PatternFill("solid", fgColor=AZUL)
txt_font = Font(name=FONTE_PADRAO, size=10)
exemplo_font = Font(name=FONTE_PADRAO, size=10, italic=True, color=CINZA_ESC)
titulo_font = Font(name=FONTE_PADRAO, size=13, bold=True, color=AZUL)
sub_font = Font(name=FONTE_PADRAO, size=10, color=CINZA_ESC)
label_font = Font(name=FONTE_PADRAO, size=10, bold=True)
input_fill = PatternFill("solid", fgColor=AMARELO)
bloco_fill = PatternFill("solid", fgColor=CINZA)
borda = Border(*[Side(style="thin", color="D3D1C7")] * 4)


def flat(wb, nome, headers, larguras=None, exemplo=None, congelar="A2"):
    """Aba plana: cabecalho na linha 1, dados a partir da linha 2."""
    ws = wb.create_sheet(nome)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = borda
    ws.row_dimensions[1].height = 22
    for i, w in enumerate(larguras or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if exemplo:
        for i, v in enumerate(exemplo, start=1):
            c = ws.cell(row=2, column=i, value=v)
            c.font = exemplo_font
    ws.freeze_panes = congelar
    ws.sheet_view.showGridLines = False
    return ws


def kv(ws, row, chave, valor, ajuda="", editavel=True):
    a = ws.cell(row=row, column=1, value=chave)
    a.font = label_font
    b = ws.cell(row=row, column=3, value=valor)
    b.font = txt_font
    if editavel:
        b.fill = input_fill
    b.border = borda
    d = ws.cell(row=row, column=4, value=ajuda)
    d.font = sub_font
    return row + 1


def build(saida):
    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ INICIO
    ws = wb.create_sheet("INÍCIO")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 95
    ws["B2"] = "GSL-DADOS — camada restrita do sistema GSL Bartofil"
    ws["B2"].font = titulo_font
    ws["B3"] = ("Esta planilha concentra os dados individuais. Ela NAO deve ser compartilhada "
                "com os coordenadores.")
    ws["B3"].font = sub_font

    linhas = [
        ("", ""),
        ("O QUE EDITAR", "Somente as celulas amarelas e as abas DIM_ e FONTES/PARAM."),
        ("", "As abas STG_, FATOS, PENDENCIAS, SYNC e LOG sao escritas pelo script."),
        ("", ""),
        ("1 · PARAM", "IDs, e-mails e parametros gerais. Preencha antes de qualquer coisa."),
        ("2 · FONTES", "ID e aba de cada planilha de origem + de-para dos cabecalhos."),
        ("3 · DIM_PESSOAS", "Matricula, nome oficial e apelidos (grafias ja vistas)."),
        ("4 · DIM_TURNO_HIST", "Turno com vigencia. A falta pertence ao turno da data do fato."),
        ("5 · DIM_CODIGOS", "Cada codigo das planilhas de origem -> tipo canonico."),
        ("6 · DIM_SETORES", "Nome do setor na origem -> nome oficial do CD."),
        ("", ""),
        ("ROTINA", "Menu GSL-DADOS > Importar agora. Automatico de hora em hora, 6h-20h."),
        ("PENDENCIAS", "O que o script nao traduziu cai la. Resolva toda semana."),
        ("PUBLICACAO", "Diaria as 5h30: envia SOMENTE agregados para a planilha do GSL."),
        ("", ""),
        ("PRIVACIDADE", "Nao registre CID, diagnostico ou imagem de atestado. Nunca."),
        ("RETENCAO", "Dado individual: 5 anos. Depois, apenas o agregado historico."),
    ]
    r = 5
    for a, b in linhas:
        if a:
            c = ws.cell(row=r, column=2, value=a)
            c.font = label_font
        c = ws.cell(row=r, column=3, value=b)
        c.font = txt_font
        r += 1

    # ------------------------------------------------------------------- PARAM
    ws = wb.create_sheet("PARAM")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 70
    ws["A1"] = "PARAM — parametros do sistema"
    ws["A1"].font = titulo_font
    ws["A2"] = "Celulas amarelas: preencher. O script localiza cada linha pelo texto da coluna A."
    ws["A2"].font = sub_font
    r = 4
    r = kv(ws, r, "ID da planilha do GSL (destino)", "", "ID do calendario que os coordenadores usam")
    r = kv(ws, r, "E-mail do administrador", "hercules@bartofil.com.br", "recebe falhas e pendencias")
    r = kv(ws, r, "E-mail do gerente", "gerente@bartofil.com.br", "")
    r = kv(ws, r, "E-mail do RH", "rh@bartofil.com.br", "recebe pendencias de pessoas e codigos")
    r = kv(ws, r, "Janela de leitura (dias)", 60, "quantos dias para tras cada importacao le")
    r = kv(ws, r, "Limiar de similaridade de nomes", 0.88, "0 a 1. Abaixo disso vai para PENDENCIAS")
    r = kv(ws, r, "Publicar agregados no GSL (S/N)", "S", "N pausa a publicacao sem desligar o ETL")
    r = kv(ws, r, "Aba de destino no GSL", "OCORRÊNCIAS", "nao mudar sem ajustar o Code.gs do GSL")
    r = kv(ws, r, "Primeira linha de dados no GSL", 6, "cabecalho do GSL fica na linha 5")
    r = kv(ws, r, "Contar falta justificada no score (S/N)", "N", "decisao da gerencia")
    r = kv(ws, r, "Peso da saida antecipada (em faltas)", 0.5, "0 = nao entra no absenteismo")

    # ------------------------------------------------------------------ FONTES
    ws = wb.create_sheet("FONTES")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [16, 12, 46, 22, 18, 14, 60]):
        ws.column_dimensions[col].width = w
    ws["A1"] = "FONTES — as tres planilhas da empresa"
    ws["A1"].font = titulo_font
    ws["A2"] = ("Bloco 1: onde esta cada planilha. Bloco 2: como o cabecalho dela se chama. "
                "O script procura os blocos pelos titulos das colunas, entao pode inserir linhas.")
    ws["A2"].font = sub_font

    ws["A4"] = "BLOCO 1 — LOCALIZACAO"
    ws["A4"].font = label_font
    ws["A4"].fill = bloco_fill
    hb1 = ["FONTE", "ATIVA", "ID DA PLANILHA", "ABA", "LINHA DO CABEÇALHO", "JANELA (DIAS)", "OBSERVAÇÃO"]
    for i, h in enumerate(hb1, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
    for i, f in enumerate(["RH", "ERROS", "METAS"]):
        row = 6 + i
        ws.cell(row=row, column=1, value=f).font = label_font
        for col, val in [(2, "S"), (3, ""), (4, ""), (5, 1), (6, "")]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = txt_font
            c.fill = input_fill
            c.border = borda
        ws.cell(row=row, column=7, value="").font = txt_font

    ws["A11"] = "BLOCO 2 — DE-PARA DE CABEÇALHOS"
    ws["A11"].font = label_font
    ws["A11"].fill = bloco_fill
    ws["A12"] = ("Escreva o nome EXATO da coluna na planilha de origem. Varias grafias: separe por ; "
                 "Deixe em branco o campo que a origem nao tem.")
    ws["A12"].font = sub_font
    hb2 = ["FONTE", "CAMPO", "CABEÇALHO NA ORIGEM", "OBRIGATÓRIO", "OBSERVAÇÃO"]
    for i, h in enumerate(hb2, start=1):
        c = ws.cell(row=13, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill

    mapa = [
        ("RH", "data", "DATA;DIA;DATA DA OCORRENCIA", "S", "data do fato"),
        ("RH", "matricula", "MATRICULA;CHAPA;REGISTRO", "N", "chave ideal — peça ao RH se não existir"),
        ("RH", "colaborador", "COLABORADOR;NOME;FUNCIONARIO", "N", "usado se não houver matrícula"),
        ("RH", "turno", "TURNO;ESCALA", "N", "se vazio, vem do DIM_TURNO_HIST"),
        ("RH", "codigo", "CODIGO;COD;OCORRENCIA;TIPO", "S", "traduzido em DIM_CODIGOS"),
        ("RH", "qtde", "QTDE;QUANTIDADE;DIAS;HORAS", "N", "vazio = 1"),
        ("RH", "setor", "SETOR;AREA", "N", ""),
        ("RH", "descricao", "OBSERVACAO;OBS;MOTIVO", "N", "nunca colar CID ou diagnóstico"),
        ("ERROS", "data", "DATA;DIA", "S", ""),
        ("ERROS", "matricula", "MATRICULA;CHAPA", "N", ""),
        ("ERROS", "colaborador", "COLABORADOR;NOME;RESPONSAVEL", "N", ""),
        ("ERROS", "turno", "TURNO", "N", ""),
        ("ERROS", "setor", "SETOR;AREA;PROCESSO", "S", ""),
        ("ERROS", "codigo", "CODIGO;TIPO DE ERRO;OCORRENCIA", "N", "vazio = Erro operacional"),
        ("ERROS", "qtde", "QTDE;QUANTIDADE;OCORRENCIAS", "N", "vazio = 1"),
        ("ERROS", "descricao", "DESCRICAO;CAUSA;OBSERVACAO", "N", ""),
        ("METAS", "mes", "MES;COMPETENCIA;PERIODO", "S", "qualquer data do mês serve"),
        ("METAS", "setor", "SETOR;AREA", "S", ""),
        ("METAS", "turno", "TURNO", "N", "vazio = meta do setor inteiro"),
        ("METAS", "indicador", "INDICADOR;META;KPI", "S", "nome do que é medido"),
        ("METAS", "meta", "VALOR DA META;OBJETIVO;META", "S", ""),
        ("METAS", "realizado", "REALIZADO;RESULTADO;ATINGIDO", "S", ""),
    ]
    for i, (fonte, campo, cab, obg, obs) in enumerate(mapa):
        row = 14 + i
        ws.cell(row=row, column=1, value=fonte).font = txt_font
        ws.cell(row=row, column=2, value=campo).font = label_font
        c = ws.cell(row=row, column=3, value=cab)
        c.font = txt_font
        c.fill = input_fill
        c.border = borda
        ws.cell(row=row, column=4, value=obg).font = txt_font
        ws.cell(row=row, column=5, value=obs).font = sub_font
    ws.freeze_panes = "A14"

    # -------------------------------------------------------------- DIMENSOES
    flat(wb, "DIM_PESSOAS",
         ["MATRICULA", "NOME OFICIAL", "APELIDOS", "SETOR", "ADMISSAO", "DESLIGAMENTO", "ATIVO"],
         [14, 34, 46, 26, 14, 14, 10],
         ["00123", "JOÃO DA SILVA (exemplo — apague)", "JOAO SILVA;J. SILVA;JOÃO S.",
          "Separação (Picking)", "2024-03-11", "", "S"])

    flat(wb, "DIM_TURNO_HIST",
         ["MATRICULA", "TURNO", "VALIDO DE", "VALIDO ATE", "OBSERVACAO"],
         [14, 10, 14, 14, 46],
         ["00123", "A", "2024-03-11", "", "(exemplo — apague) vazio em VALIDO ATE = vigente"])

    ws = flat(wb, "DIM_CODIGOS",
              ["ORIGEM", "CODIGO", "DESCRICAO NA ORIGEM", "TIPO CANONICO", "UNIDADE",
               "CONTA NO SCORE", "GRAVIDADE"],
              [12, 12, 40, 24, 12, 16, 12],
              ["RH", "F", "FALTA NAO JUSTIFICADA (exemplo — apague)", "Falta", "evento", "S", "3"])
    tipos = ("\"Falta,Falta justificada,Atestado,Saída antecipada,Atraso,"
             "Erro operacional,Férias,Afastamento,Ignorar\"")
    dv = DataValidation(type="list", formula1=tipos, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("D2:D500")
    dv2 = DataValidation(type="list", formula1="\"evento,dia,hora\"", allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv2)
    dv2.add("E2:E500")
    dv3 = DataValidation(type="list", formula1="\"S,N\"", allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv3)
    dv3.add("F2:F500")

    flat(wb, "DIM_SETORES",
         ["NOME NA ORIGEM", "SETOR OFICIAL", "OBSERVACAO"],
         [36, 36, 46],
         ["PICKING (exemplo — apague)", "Separação (Picking)", "as grafias da origem entram aqui"])

    # ---------------------------------------------------------------- STAGING
    stg = ["_ORIGEM", "_LINHA_ORIG", "_HASH", "_IMPORTADO_EM",
           "DATA", "MATRICULA", "COLABORADOR", "TURNO", "SETOR", "CODIGO", "QTDE", "DESCRICAO"]
    for nome in ("STG_RH", "STG_ERROS"):
        flat(wb, nome, stg, [12, 14, 30, 20, 14, 14, 30, 10, 26, 14, 10, 40])
    flat(wb, "STG_METAS",
         ["_ORIGEM", "_LINHA_ORIG", "_HASH", "_IMPORTADO_EM",
          "MES", "SETOR", "TURNO", "INDICADOR", "META", "REALIZADO"],
         [12, 14, 30, 20, 14, 26, 10, 30, 14, 14])

    # ------------------------------------------------------------------ FATOS
    flat(wb, "FATOS",
         ["DATA", "TIPO", "TURNO", "SETOR", "QTDE", "MATRICULA", "COLABORADOR", "DESCRICAO",
          "SEMANA", "MES", "ANO", "_ORIGEM", "_LINHA_ORIG", "_HASH", "_IMPORTADO_EM"],
         [12, 20, 8, 26, 8, 14, 30, 40, 10, 12, 8, 12, 14, 30, 20])

    flat(wb, "FATOS_METAS",
         ["MES", "SETOR", "TURNO", "INDICADOR", "META", "REALIZADO", "PCT_ATINGIDO",
          "_ORIGEM", "_LINHA_ORIG", "_HASH", "_IMPORTADO_EM"],
         [12, 26, 8, 30, 12, 12, 14, 12, 14, 30, 20])

    # ------------------------------------------------------------- PENDENCIAS
    ws = flat(wb, "PENDENCIAS",
              ["_ORIGEM", "_LINHA_ORIG", "TIPO", "VALOR NAO RECONHECIDO", "SUGESTAO",
               "SIMILARIDADE", "ACAO", "RESOLVIDO EM"],
              [12, 14, 14, 40, 40, 14, 22, 20],
              ["RH", 42, "pessoa", "JOAO DA SILVAA (exemplo — apague)", "JOÃO DA SILVA", 0.94, "", ""])
    dv = DataValidation(type="list",
                        formula1="\"Confirmar sugestão,Criar novo,Ignorar\"",
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("G2:G2000")

    # -------------------------------------------------------------- SYNC e LOG
    flat(wb, "SYNC",
         ["FONTE", "ULTIMA LEITURA", "LINHAS LIDAS", "FATOS GERADOS", "IGNORADAS",
          "PENDENCIAS", "STATUS", "MENSAGEM"],
         [12, 20, 14, 16, 12, 14, 14, 70])

    flat(wb, "LOG",
         ["QUANDO", "QUEM", "ACAO", "ALVO", "VALOR ANTERIOR", "VALOR NOVO", "DETALHE"],
         [20, 28, 22, 26, 30, 30, 50])

    wb.save(saida)
    return saida


if __name__ == "__main__":
    destino = sys.argv[1] if len(sys.argv) > 1 else "GSL-DADOS-v7.xlsx"
    print(build(destino))
